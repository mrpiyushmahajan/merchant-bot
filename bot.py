"""
Merchant Turnover Telegram Bot
================================
Accepts GPay Business & Paytm merchant CSVs, compiles them into a
colour-coded Excel report with daily turnovers and a grand total.

Commands
--------
/start   – welcome & instructions
/status  – show what has been collected so far
/compile – generate and send the Excel report
/reset   – clear all data for a fresh run (asks confirmation)
"""

import io
import logging
import os
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import pandas as pd
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from telegram import InlineKeyboardButton, InlineKeyboardMarkup, Update
from telegram.ext import (
    Application,
    CallbackQueryHandler,
    CommandHandler,
    ContextTypes,
    MessageHandler,
    PicklePersistence,
    filters,
)

load_dotenv()
BOT_TOKEN = os.getenv("BOT_TOKEN", "")

logging.basicConfig(
    format="%(asctime)s  %(levelname)-8s  %(name)s  %(message)s",
    level=logging.INFO,
)
logger = logging.getLogger(__name__)


# ══════════════════════════════════════════════════════════════════════════════
#  CSV DETECTION & PARSING
# ══════════════════════════════════════════════════════════════════════════════

def _clean(val: str) -> str:
    """Strip whitespace and stray single/double quotes from a cell value."""
    return str(val).strip().strip("'\"")


def _find_col(df: pd.DataFrame, *candidates: str):
    """
    Return the actual DataFrame column name that best matches any candidate.
    Tries exact match (case-insensitive) first, then prefix match.
    Returns None if nothing found.
    """
    cols_lower = {c.strip().lower(): c.strip() for c in df.columns}
    for cand in candidates:
        cand_l = cand.lower()
        if cand_l in cols_lower:
            return cols_lower[cand_l]
        # prefix match — e.g. "type" matches "Type (UPI / UPI CC)"
        for orig_l, orig in cols_lower.items():
            if orig_l.startswith(cand_l):
                return orig
    return None


def detect_csv_type(df: pd.DataFrame) -> str:
    """Return 'gpay', 'paytm', or 'unknown' based on column signatures."""
    cols = {c.strip().lower() for c in df.columns}
    if "transaction_date" in cols or "merchant_name" in cols:
        return "paytm"
    if "creation time" in cols or "paid via" in cols:
        return "gpay"
    return "unknown"


def parse_gpay_csv(df: pd.DataFrame, shop_name: str) -> list[dict]:
    """
    GPay Business export — works with both old and new column layouts:

    Old layout : Type | Payer/Receiver | Creation time | Status | Amount
    New layout : Type (UPI / UPI CC) | Payer | Creation time | Status | Amount

    Rules:
    • Type    starts with 'UPI'
    • Status  is 'Settled' OR 'Scheduled to Settle'
    • Amount  > 0
    """
    df = df.copy()
    df.columns = [c.strip() for c in df.columns]

    type_col   = _find_col(df, "type")
    status_col = _find_col(df, "status")
    amount_col = _find_col(df, "amount")
    time_col   = _find_col(df, "creation time")
    payer_col  = _find_col(df, "payer/receiver", "payer")
    txn_col    = _find_col(df, "transaction id")

    if not all([type_col, status_col, amount_col, time_col]):
        logger.warning("GPay CSV missing expected columns. Found: %s", list(df.columns))
        return []

    df[amount_col] = pd.to_numeric(df[amount_col], errors="coerce")
    status_lower   = df[status_col].str.strip().str.lower()
    type_upper     = df[type_col].str.strip().str.upper()

    mask = (
        (status_lower.str.startswith("settled") | status_lower.str.startswith("scheduled to settle"))
        & type_upper.str.startswith("UPI")
        & (df[amount_col] > 0)
    )
    df = df[mask]

    records: list[dict] = []
    for _, row in df.iterrows():
        try:
            raw_time = _clean(row[time_col])
            date = pd.to_datetime(raw_time, dayfirst=True).date()
            records.append({
                "shop":     shop_name,
                "merchant": "GPay",
                "date":     date,
                "amount":   float(row[amount_col]),
                "payer":    _clean(row[payer_col]) if payer_col else "",
                "txn_id":   _clean(row[txn_col])   if txn_col   else "",
            })
        except Exception as exc:
            logger.warning("Skipping GPay row – %s", exc)

    return records


def parse_paytm_csv(df: pd.DataFrame) -> list[dict]:
    """
    Paytm merchant export — reads Merchant_Name from the CSV.

    Rules:
    • Status == 'SUCCESS'
    • Amount >  0
    """
    df = df.copy()
    df.columns = [c.strip() for c in df.columns]

    df["_status"] = df["Status"].apply(_clean)
    df = df[df["_status"] == "SUCCESS"]

    df["Amount"] = pd.to_numeric(
        df["Amount"].apply(lambda x: _clean(x).replace(",", "")),
        errors="coerce",
    )
    df = df[df["Amount"] > 0]

    records: list[dict] = []
    for _, row in df.iterrows():
        try:
            date = pd.to_datetime(_clean(row["Transaction_Date"])).date()
            shop = _clean(row.get("Merchant_Name", "Unknown"))
            records.append({
                "shop":     shop,
                "merchant": "Paytm",
                "date":     date,
                "amount":   float(row["Amount"]),
                "payer":    _clean(row.get("Customer_VPA", "")),
                "txn_id":   _clean(row.get("Transaction_ID", "")),
            })
        except Exception as exc:
            logger.warning("Skipping Paytm row – %s", exc)

    return records


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL REPORT GENERATION
# ══════════════════════════════════════════════════════════════════════════════

C = {
    "header_bg":      "263238",
    "title_bg":       "1A237E",
    "gpay_row":       "E3F2FD",
    "gpay_accent":    "1565C0",
    "paytm_row":      "E8F5E9",
    "paytm_accent":   "2E7D32",
    "subtotal_bg":    "FFF9C4",
    "subtotal_font":  "E65100",
    "grand_bg":       "BF360C",
    "white":          "FFFFFF",
    "grey_border":    "B0BEC5",
}

_thin  = Side(style="thin",   color=C["grey_border"])
_thick = Side(style="medium", color="78909C")


def _border(top=_thin, bottom=_thin, left=_thin, right=_thin):
    return Border(top=top, bottom=bottom, left=left, right=right)


def _hdr(ws, row, col, value, bg="header_bg", fg="white", size=10,
         bold=True, h_align="center"):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = Font(bold=bold, color=C[fg], size=size, name="Calibri")
    cell.fill      = PatternFill("solid", fgColor=C[bg])
    cell.alignment = Alignment(horizontal=h_align, vertical="center", wrap_text=True)


def _data(ws, row, col, value, bg, bold=False, fmt=None, fg="header_bg",
          h_align="center"):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = Font(bold=bold, color=C[fg], name="Calibri", size=10)
    cell.fill      = PatternFill("solid", fgColor=C[bg])
    cell.alignment = Alignment(horizontal=h_align, vertical="center")
    if fmt:
        cell.number_format = fmt


def _set_widths(ws, widths):
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w


def _freeze(ws, cell):
    ws.freeze_panes = cell


def _build_dashboard(wb, shop_daily, shop_txns):
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    cell = ws["A1"]
    cell.value = (
        f"  Merchant Daily Turnover Report  •  "
        f"Generated {datetime.now().strftime('%d %b %Y  %H:%M')}"
    )
    cell.font      = Font(bold=True, size=13, color=C["white"], name="Calibri")
    cell.fill      = PatternFill("solid", fgColor=C["title_bg"])
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 32

    col_labels = ["Shop Name", "Merchant", "Date", "Transactions", "Daily Turnover (₹)", ""]
    for c, lbl in enumerate(col_labels, 1):
        _hdr(ws, 2, c, lbl)
        ws.cell(row=2, column=c).border = _border(
            top=Side(style="medium", color="455A64"),
            bottom=Side(style="medium", color="455A64"),
        )
    ws.row_dimensions[2].height = 22

    row = 3
    grand_total = 0.0

    for (shop, merchant), daily in sorted(shop_daily.items(), key=lambda x: x[0][0]):
        row_bg     = "gpay_row"    if merchant == "GPay" else "paytm_row"
        accent_key = "gpay_accent" if merchant == "GPay" else "paytm_accent"
        accent_hex = C[accent_key]

        shop_total  = sum(daily.values())
        grand_total += shop_total
        txns_total  = len(shop_txns[(shop, merchant)])
        first_row   = row

        for date in sorted(daily.keys()):
            day_txns = sum(1 for r in shop_txns[(shop, merchant)] if r["date"] == date)
            day_amt  = daily[date]
            is_first = (row == first_row)

            name_cell = ws.cell(row=row, column=1, value=shop if is_first else "")
            name_cell.fill      = PatternFill("solid", fgColor=C[row_bg])
            name_cell.font      = Font(bold=is_first, color=accent_hex, name="Calibri", size=10)
            name_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

            mc = ws.cell(row=row, column=2, value=merchant if is_first else "")
            mc.fill      = PatternFill("solid", fgColor=C[row_bg])
            mc.font      = Font(bold=True, color=accent_hex, name="Calibri", size=10)
            mc.alignment = Alignment(horizontal="center", vertical="center")

            _data(ws, row, 3, date.strftime("%d %b %Y"), row_bg, h_align="center")
            _data(ws, row, 4, day_txns,  row_bg, h_align="center")
            _data(ws, row, 5, day_amt,   row_bg, fmt="#,##0.00", h_align="right")

            for c in range(1, 6):
                ws.cell(row=row, column=c).border = _border()
            ws.row_dimensions[row].height = 18
            row += 1

        ws.merge_cells(f"A{row}:B{row}")
        sub = ws.cell(row=row, column=1, value=f"  {shop} — Subtotal")
        sub.font      = Font(bold=True, color=C["subtotal_font"], name="Calibri", size=10)
        sub.fill      = PatternFill("solid", fgColor=C["subtotal_bg"])
        sub.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        _data(ws, row, 3, "", "subtotal_bg")
        _data(ws, row, 4, txns_total, "subtotal_bg", bold=True, fg="subtotal_font", h_align="center")
        _data(ws, row, 5, shop_total, "subtotal_bg", bold=True, fg="subtotal_font", fmt="#,##0.00", h_align="right")
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = _border(
                top=Side(style="medium", color="E65100"),
                bottom=Side(style="medium", color="E65100"),
            )
        ws.row_dimensions[row].height = 20
        row += 1

    ws.merge_cells(f"A{row}:D{row}")
    gt = ws.cell(row=row, column=1, value="  GRAND TOTAL — All Shops")
    gt.font      = Font(bold=True, size=12, color=C["white"], name="Calibri")
    gt.fill      = PatternFill("solid", fgColor=C["grand_bg"])
    gt.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for c in range(2, 5):
        ws.cell(row=row, column=c).fill = PatternFill("solid", fgColor=C["grand_bg"])
    gv = ws.cell(row=row, column=5, value=grand_total)
    gv.font          = Font(bold=True, size=12, color=C["white"], name="Calibri")
    gv.fill          = PatternFill("solid", fgColor=C["grand_bg"])
    gv.number_format = "#,##0.00"
    gv.alignment     = Alignment(horizontal="right", vertical="center")
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = _border(
            top=Side(style="thick", color="BF360C"),
            bottom=Side(style="thick", color="BF360C"),
        )
    ws.row_dimensions[row].height = 26

    _set_widths(ws, {"A": 24, "B": 12, "C": 16, "D": 16, "E": 22, "F": 4})
    _freeze(ws, "A3")


def _build_shop_summary(wb, shop_daily, shop_txns):
    ws = wb.create_sheet("Shop Summary")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    cell = ws["A1"]
    cell.value     = "  Shop-wise Summary"
    cell.font      = Font(bold=True, size=13, color=C["white"], name="Calibri")
    cell.fill      = PatternFill("solid", fgColor=C["title_bg"])
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    hdrs = ["#", "Shop Name", "Merchant", "Total Transactions",
            "Total Turnover (₹)", "Date From", "Date To", ""]
    for c, h in enumerate(hdrs, 1):
        _hdr(ws, 2, c, h)
    ws.row_dimensions[2].height = 22

    idx = 1
    grand = 0.0
    for (shop, merchant), daily in sorted(shop_daily.items(), key=lambda x: x[0][0]):
        row_bg     = "gpay_row"    if merchant == "GPay" else "paytm_row"
        accent_key = "gpay_accent" if merchant == "GPay" else "paytm_accent"
        total      = sum(daily.values())
        grand     += total
        txns       = len(shop_txns[(shop, merchant)])
        dates      = sorted(daily.keys())
        r          = idx + 2

        _data(ws, r, 1, idx,    row_bg, h_align="center")
        _data(ws, r, 2, shop,   row_bg, bold=True, fg=accent_key, h_align="left")
        _data(ws, r, 3, merchant, row_bg, bold=True, fg=accent_key)
        _data(ws, r, 4, txns,   row_bg, h_align="center")
        _data(ws, r, 5, total,  row_bg, bold=True, fmt="#,##0.00", h_align="right")
        _data(ws, r, 6, dates[0].strftime("%d %b %Y"),  row_bg)
        _data(ws, r, 7, dates[-1].strftime("%d %b %Y"), row_bg)
        for c in range(1, 8):
            ws.cell(row=r, column=c).border = _border()
        ws.row_dimensions[r].height = 20
        idx += 1

    gt_row = idx + 2
    ws.merge_cells(f"A{gt_row}:D{gt_row}")
    gt = ws.cell(row=gt_row, column=1, value="  GRAND TOTAL")
    gt.font      = Font(bold=True, size=12, color=C["white"], name="Calibri")
    gt.fill      = PatternFill("solid", fgColor=C["grand_bg"])
    gt.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for c in range(2, 5):
        ws.cell(row=gt_row, column=c).fill = PatternFill("solid", fgColor=C["grand_bg"])
    gv = ws.cell(row=gt_row, column=5, value=grand)
    gv.font          = Font(bold=True, size=12, color=C["white"], name="Calibri")
    gv.fill          = PatternFill("solid", fgColor=C["grand_bg"])
    gv.number_format = "#,##0.00"
    gv.alignment     = Alignment(horizontal="right", vertical="center")
    for c in range(6, 8):
        ws.cell(row=gt_row, column=c).fill = PatternFill("solid", fgColor=C["grand_bg"])
    for c in range(1, 8):
        ws.cell(row=gt_row, column=c).border = _border(
            top=Side(style="thick", color="BF360C"),
            bottom=Side(style="thick", color="BF360C"),
        )
    ws.row_dimensions[gt_row].height = 26

    _set_widths(ws, {"A": 5, "B": 24, "C": 12, "D": 20, "E": 22, "F": 14, "G": 14, "H": 4})
    _freeze(ws, "A3")


def _build_all_transactions(wb, all_records):
    ws = wb.create_sheet("All Transactions")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    cell = ws["A1"]
    cell.value     = "  All Transactions"
    cell.font      = Font(bold=True, size=13, color=C["white"], name="Calibri")
    cell.fill      = PatternFill("solid", fgColor=C["title_bg"])
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    hdrs = ["#", "Shop Name", "Merchant", "Date", "Payer / VPA",
            "Transaction ID", "Amount (₹)", ""]
    for c, h in enumerate(hdrs, 1):
        _hdr(ws, 2, c, h)
    ws.row_dimensions[2].height = 22

    sorted_recs = sorted(all_records, key=lambda x: (x["shop"], x["date"]))
    for i, rec in enumerate(sorted_recs, 1):
        r          = i + 2
        row_bg     = "gpay_row"    if rec["merchant"] == "GPay" else "paytm_row"
        accent_key = "gpay_accent" if rec["merchant"] == "GPay" else "paytm_accent"

        _data(ws, r, 1, i,                              row_bg, h_align="center")
        _data(ws, r, 2, rec["shop"],                    row_bg, bold=True, fg=accent_key, h_align="left")
        _data(ws, r, 3, rec["merchant"],                row_bg, bold=True, fg=accent_key)
        _data(ws, r, 4, rec["date"].strftime("%d %b %Y"), row_bg)
        _data(ws, r, 5, rec.get("payer", ""),           row_bg, h_align="left")
        _data(ws, r, 6, rec.get("txn_id", ""),          row_bg, h_align="left")
        _data(ws, r, 7, rec["amount"],                  row_bg, fmt="#,##0.00", h_align="right")
        for c in range(1, 8):
            ws.cell(row=r, column=c).border = _border()
        ws.row_dimensions[r].height = 17

    ws.auto_filter.ref = f"A2:G{len(sorted_recs) + 2}"
    _set_widths(ws, {"A": 5, "B": 24, "C": 12, "D": 16, "E": 30, "F": 36, "G": 16, "H": 4})
    _freeze(ws, "A3")


def generate_excel(all_records: list[dict]) -> bytes:
    """Build a 3-sheet workbook from all collected records and return bytes."""
    shop_daily: dict = defaultdict(lambda: defaultdict(float))
    shop_txns:  dict = defaultdict(list)

    for rec in all_records:
        key = (rec["shop"], rec["merchant"])
        shop_daily[key][rec["date"]] += rec["amount"]
        shop_txns[key].append(rec)

    wb = Workbook()
    _build_dashboard(wb, shop_daily, shop_txns)
    _build_shop_summary(wb, shop_daily, shop_txns)
    _build_all_transactions(wb, all_records)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ══════════════════════════════════════════════════════════════════════════════
#  TELEGRAM BOT HANDLERS
# ══════════════════════════════════════════════════════════════════════════════

WELCOME = (
    "👋 *Welcome to the Merchant Turnover Bot!*\n\n"
    "I compile *GPay Business* and *Paytm* merchant CSVs from multiple shops "
    "into a single, colour-coded Excel report.\n\n"
    "📤 *How to use*\n"
    "1. Send your CSV files (one or many at once)\n"
    "2. Paytm CSVs are processed instantly — shop name is read automatically\n"
    "3. For GPay CSVs, tap your shop name or type a new one\n"
    "4. When done, tap *Compile now* or use /compile\n\n"
    "📋 *Commands*\n"
    "/status   — See what's collected so far\n"
    "/compile  — Generate & download the Excel report\n"
    "/reset    — Clear all data and start fresh\n\n"
    "Go ahead — send your first CSV file! 📁"
)

MAX_KNOWN_SHOPS = 10   # how many shop names to remember


def _ensure_state(context: ContextTypes.DEFAULT_TYPE) -> None:
    context.user_data.setdefault("records",          [])
    context.user_data.setdefault("pending_csv",      None)
    context.user_data.setdefault("gpay_queue",       [])
    context.user_data.setdefault("asking_gpay_name", False)
    context.user_data.setdefault("known_shops",      [])   # remembered shop names
    context.user_data.setdefault("seen_txn_ids",     set()) # for duplicate detection


def _add_records(context: ContextTypes.DEFAULT_TYPE,
                 new_records: list[dict]) -> tuple[list[dict], int]:
    """
    Add records to the session, skipping any whose txn_id was seen before.
    Returns (list_of_added_records, skipped_count).
    """
    seen    = context.user_data["seen_txn_ids"]
    added   = []
    skipped = 0
    for rec in new_records:
        tid = rec.get("txn_id", "")
        if tid and tid in seen:
            skipped += 1
        else:
            if tid:
                seen.add(tid)
            context.user_data["records"].append(rec)
            added.append(rec)
    return added, skipped


def _remember_shop(context: ContextTypes.DEFAULT_TYPE, shop_name: str) -> None:
    """Add shop_name to the front of known_shops, capped at MAX_KNOWN_SHOPS."""
    known = context.user_data["known_shops"]
    if shop_name in known:
        known.remove(shop_name)       # move to front
    known.insert(0, shop_name)
    del known[MAX_KNOWN_SHOPS:]       # keep list bounded


def _shop_keyboard(known_shops: list[str]) -> InlineKeyboardMarkup:
    """Inline keyboard: one button per known shop, plus a 'type new name' button."""
    rows = [[InlineKeyboardButton(s, callback_data=f"shop:{s}")] for s in known_shops]
    rows.append([InlineKeyboardButton("➕ Type a new name", callback_data="newshop")])
    return InlineKeyboardMarkup(rows)


def _compile_prompt_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([[
        InlineKeyboardButton("📊 Compile now",         callback_data="compile_now"),
        InlineKeyboardButton("⏳ Wait for more files", callback_data="wait_more"),
    ]])


def _reset_keyboard(n_records: int) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([[
        InlineKeyboardButton(f"✅ Yes, clear {n_records} record(s)", callback_data="reset_yes"),
        InlineKeyboardButton("❌ Cancel", callback_data="reset_no"),
    ]])


def _post_compile_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([[
        InlineKeyboardButton("📁 Keep data",           callback_data="keep_data"),
        InlineKeyboardButton("🔄 Clear & start fresh", callback_data="reset_yes"),
    ]])


async def _disable_buttons(query) -> None:
    """Remove the inline keyboard from the message that triggered this callback."""
    try:
        await query.edit_message_reply_markup(reply_markup=InlineKeyboardMarkup([]))
    except Exception:
        pass   # message might be too old or already edited


async def _do_compile(message, context: ContextTypes.DEFAULT_TYPE) -> bool:
    """
    Generate and send the Excel report.
    Returns True on success, False on failure.
    `message` is any telegram Message object.
    """
    records = context.user_data["records"]
    msg     = await message.reply_text("⏳ Generating your Excel report…")
    try:
        excel_bytes = generate_excel(records)
    except Exception as exc:
        logger.exception("Excel generation failed")
        await msg.edit_text(f"❌ Error generating report:\n`{exc}`", parse_mode="Markdown")
        return False

    shops       = {f"{r['shop']} ({r['merchant']})" for r in records}
    grand_total = sum(r["amount"] for r in records)
    fname       = f"Merchant_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    bio      = io.BytesIO(excel_bytes)
    bio.name = fname

    await message.reply_document(
        document=bio,
        filename=fname,
        caption=(
            f"✅ *Report ready!*\n"
            f"🏪 Shops: {len(shops)}\n"
            f"📊 Transactions: {len(records)}\n"
            f"💰 Grand Total: ₹{grand_total:,.2f}"
        ),
        parse_mode="Markdown",
    )
    await msg.delete()
    return True


async def _prompt_next_gpay(message, context: ContextTypes.DEFAULT_TYPE) -> None:
    """
    Ask the user for the shop name of the next queued GPay CSV.
    When the queue is empty, show the compile-now / wait buttons.
    `message` is any telegram Message object (from update.message or callback_query.message).
    """
    queue = context.user_data["gpay_queue"]

    if not queue:
        context.user_data["asking_gpay_name"] = False
        n = len(context.user_data["records"])
        await message.reply_text(
            f"✅ All files processed!  *{n}* transaction(s) collected.\n"
            "Send more files, or:",
            parse_mode="Markdown",
            reply_markup=_compile_prompt_keyboard(),
        )
        return

    _, filename = queue[0]
    context.user_data["asking_gpay_name"] = True
    count  = len(queue)
    known  = context.user_data.get("known_shops", [])
    plural = f"({count} GPay file(s) left)  " if count > 1 else ""

    text = (
        f"📗 *GPay CSV:* `{filename}`\n"
        f"{plural}What is the *shop name* for this file?"
    )

    if known:
        text += "\n_(Tap a name below or type a new one)_"
        await message.reply_text(text, parse_mode="Markdown",
                                 reply_markup=_shop_keyboard(known))
    else:
        text += "\n_(Type the shop name and send)_"
        await message.reply_text(text, parse_mode="Markdown")


async def _process_gpay_with_name(message, context: ContextTypes.DEFAULT_TYPE,
                                   shop_name: str) -> None:
    """
    Pop the first GPay CSV from the queue, parse it with shop_name,
    deduplicate, add records, then prompt for the next file (or compile).
    """
    queue = context.user_data["gpay_queue"]
    if not queue:
        context.user_data["asking_gpay_name"] = False
        return

    _remember_shop(context, shop_name)

    csv_bytes, filename = queue.pop(0)
    df          = pd.read_csv(io.BytesIO(csv_bytes))
    new_records = parse_gpay_csv(df, shop_name)
    added, skipped = _add_records(context, new_records)

    if new_records:
        added_total  = sum(r["amount"] for r in added)
        skip_note    = f"\n⚠️ {skipped} duplicate txn(s) skipped" if skipped else ""
        session_n    = len(context.user_data["records"])
        await message.reply_text(
            f"✅ *GPay processed:* `{filename}`\n"
            f"🏪 Shop: `{shop_name}`  •  📊 {len(added)} txns  •  💰 ₹{added_total:,.2f}"
            f"{skip_note}\n"
            f"📦 Session total: {session_n} records",
            parse_mode="Markdown",
        )
    else:
        # Show diagnostic info
        try:
            df_check = pd.read_csv(io.BytesIO(csv_bytes))
            df_check.columns = [c.strip() for c in df_check.columns]
            _a = _find_col(df_check, "amount")
            _t = _find_col(df_check, "type")
            _s = _find_col(df_check, "status")
            if _a and _t and _s:
                df_check[_a] = pd.to_numeric(df_check[_a], errors="coerce")
                _sl = df_check[_s].str.strip().str.lower()
                settled = int((_sl.str.startswith("settled") | _sl.str.startswith("scheduled to settle")).sum())
                upi     = int(df_check[_t].str.strip().str.upper().str.startswith("UPI").sum())
                pos_amt = int((df_check[_a] > 0).sum())
                detail  = f"Settled/Scheduled rows: {settled}  •  UPI rows: {upi}  •  Positive amount rows: {pos_amt}"
            else:
                detail = f"Unexpected columns: {', '.join(df_check.columns[:6])}"
        except Exception as exc:
            detail = f"Could not inspect: {exc}"
        await message.reply_text(
            f"⚠️ No valid transactions from `{filename}` for *{shop_name}*\n"
            f"{detail}\n\n"
            "Expected: UPI type + Settled (or Scheduled to Settle) + Amount > 0\n"
            "Make sure this is a GPay Business export.",
            parse_mode="Markdown",
        )

    await _prompt_next_gpay(message, context)


# ── Command handlers ───────────────────────────────────────────────────────────

async def cmd_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    _ensure_state(context)
    await update.message.reply_text(WELCOME, parse_mode="Markdown")


async def cmd_status(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    _ensure_state(context)
    records = context.user_data["records"]

    if not records:
        await update.message.reply_text(
            "📭 Nothing collected yet.  Send some CSV files to get started!"
        )
        return

    shop_totals: dict = defaultdict(float)
    shop_counts: dict = defaultdict(int)
    for r in records:
        key = f"{r['shop']}  ({r['merchant']})"
        shop_totals[key] += r["amount"]
        shop_counts[key] += 1

    lines = ["📊 *Current data:*\n"]
    for key in sorted(shop_totals):
        lines.append(
            f"🏪 *{key}*\n"
            f"   {shop_counts[key]} transactions  •  ₹{shop_totals[key]:,.2f}"
        )
    lines.append(
        f"\n💰 *Grand Total: ₹{sum(shop_totals.values()):,.2f}*\n\n"
        "Use /compile when ready."
    )
    await update.message.reply_text("\n".join(lines), parse_mode="Markdown")


async def cmd_compile(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    _ensure_state(context)
    records = context.user_data["records"]

    if context.user_data["gpay_queue"]:
        n = len(context.user_data["gpay_queue"])
        await update.message.reply_text(
            f"⚠️ {n} GPay CSV(s) still need a shop name.\n"
            "Reply with the shop name to continue, then use /compile."
        )
        return

    if not records:
        await update.message.reply_text(
            "📭 Nothing to compile yet.  Send some CSV files first!"
        )
        return

    ok = await _do_compile(update.message, context)
    if ok:
        await update.message.reply_text(
            "Keep the current data or start fresh for the next batch?",
            reply_markup=_post_compile_keyboard(),
        )


async def cmd_reset(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    _ensure_state(context)
    n = len(context.user_data["records"])
    if n == 0:
        # Nothing to lose — just clear and confirm
        context.user_data["records"]          = []
        context.user_data["pending_csv"]      = None
        context.user_data["gpay_queue"]       = []
        context.user_data["asking_gpay_name"] = False
        context.user_data["seen_txn_ids"]     = set()
        await update.message.reply_text(
            "🔄 Already empty!  Shop name history is kept.\n"
            "Send CSVs whenever you're ready."
        )
        return

    await update.message.reply_text(
        f"⚠️ This will clear *{n} transaction record(s)*.\n"
        "Your shop name history will be kept.\n\nAre you sure?",
        parse_mode="Markdown",
        reply_markup=_reset_keyboard(n),
    )


async def cmd_debug(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Hidden command to inspect current session state."""
    _ensure_state(context)
    records    = context.user_data.get("records", [])
    queue      = context.user_data.get("gpay_queue", [])
    asking     = context.user_data.get("asking_gpay_name", False)
    known      = context.user_data.get("known_shops", [])
    seen_ids   = context.user_data.get("seen_txn_ids", set())
    await update.message.reply_text(
        f"🔍 *Debug state*\n"
        f"Records collected: {len(records)}\n"
        f"GPay queue length: {len(queue)}\n"
        f"Waiting for name: {asking}\n"
        f"Known shops: {known}\n"
        f"Seen txn IDs: {len(seen_ids)}\n"
        f"Shops in records: {sorted({r['shop'] for r in records}) if records else 'none'}",
        parse_mode="Markdown",
    )


# ── Document & text handlers ───────────────────────────────────────────────────

async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    _ensure_state(context)
    doc = update.message.document

    if not doc.file_name.lower().endswith(".csv"):
        await update.message.reply_text(
            "⚠️ That doesn't look like a CSV file.  "
            "Please send a `.csv` export from GPay or Paytm."
        )
        return

    tg_file = await doc.get_file()
    bio     = io.BytesIO()
    await tg_file.download_to_memory(bio)
    bio.seek(0)

    try:
        df = pd.read_csv(bio)
    except Exception as exc:
        await update.message.reply_text(
            f"❌ Could not read `{doc.file_name}`: `{exc}`",
            parse_mode="Markdown",
        )
        return

    csv_type     = detect_csv_type(df)
    cols_preview = ", ".join(list(df.columns)[:6])

    if csv_type == "gpay":
        # Estimate how many rows will pass the filter (for display only)
        df_tmp   = df.copy()
        df_tmp.columns = [c.strip() for c in df_tmp.columns]
        _amt_col = _find_col(df_tmp, "amount")
        _typ_col = _find_col(df_tmp, "type")
        _sta_col = _find_col(df_tmp, "status")
        if _amt_col and _typ_col and _sta_col:
            df_tmp[_amt_col] = pd.to_numeric(df_tmp[_amt_col], errors="coerce")
            _sta_l   = df_tmp[_sta_col].str.strip().str.lower()
            _typ_u   = df_tmp[_typ_col].str.strip().str.upper()
            passable = int((
                (_sta_l.str.startswith("settled") | _sta_l.str.startswith("scheduled to settle"))
                & _typ_u.str.startswith("UPI")
                & (df_tmp[_amt_col] > 0)
            ).sum())
        else:
            passable = "?"

        bio.seek(0)
        context.user_data["gpay_queue"].append((bio.read(), doc.file_name))
        await update.message.reply_text(
            f"📗 *GPay CSV queued:* `{doc.file_name}`\n"
            f"📋 {len(df)} rows total  •  ~{passable} will be counted after filtering",
            parse_mode="Markdown",
        )
        if not context.user_data["asking_gpay_name"]:
            await _prompt_next_gpay(update.message, context)

    elif csv_type == "paytm":
        df_tmp = df.copy()
        df_tmp.columns = [c.strip() for c in df_tmp.columns]
        df_tmp["_s"]   = df_tmp["Status"].apply(_clean)
        total_rows    = len(df_tmp)
        success_rows  = int((df_tmp["_s"] == "SUCCESS").sum())

        new_records = parse_paytm_csv(df)
        if not new_records:
            await update.message.reply_text(
                f"⚠️ *No records kept from* `{doc.file_name}`\n"
                f"Total rows: {total_rows}  •  SUCCESS rows: {success_rows}\n\n"
                f"First columns seen: `{cols_preview}`\n\n"
                "Expected Paytm columns: `Transaction_Date`, `Merchant_Name`, `Status`, `Amount`\n"
                "Make sure this is the *merchant transaction export* from Paytm for Business.",
                parse_mode="Markdown",
            )
            return

        added, skipped = _add_records(context, new_records)
        shops     = sorted({r["shop"] for r in added})
        total     = sum(r["amount"] for r in added)
        skip_note = f"\n⚠️ {skipped} duplicate txn(s) skipped" if skipped else ""
        session_n = len(context.user_data["records"])
        await update.message.reply_text(
            f"✅ *Paytm processed:* `{doc.file_name}`\n"
            f"🏪 Shop(s): `{'`, `'.join(shops)}`\n"
            f"📊 {len(added)} transactions  •  💰 ₹{total:,.2f}"
            f"{skip_note}\n"
            f"📦 Session total so far: {session_n} records",
            parse_mode="Markdown",
        )

    else:
        await update.message.reply_text(
            f"❓ *Could not identify:* `{doc.file_name}`\n\n"
            f"Columns found: `{cols_preview}...`\n\n"
            "*GPay export* must have: `Creation time`, `Paid via`\n"
            "*Paytm export* must have: `Transaction_Date`, `Merchant_Name`\n\n"
            "Make sure you export directly from the GPay Business / Paytm for Business app.",
            parse_mode="Markdown",
        )


async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    _ensure_state(context)

    if not context.user_data["asking_gpay_name"]:
        await update.message.reply_text(
            "Send me CSV files, or use /start to see all commands."
        )
        return

    shop_name = update.message.text.strip()
    if not shop_name:
        await update.message.reply_text("Please enter a valid shop name (can't be empty).")
        return

    await _process_gpay_with_name(update.message, context, shop_name)


# ── Inline-button callback handler ────────────────────────────────────────────

async def handle_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    _ensure_state(context)
    query = update.callback_query
    await query.answer()          # dismiss the loading spinner on the button
    await _disable_buttons(query) # remove buttons so they can't be tapped twice
    data  = query.data
    msg   = query.message         # the message the buttons were attached to

    # ── Shop name button tapped ──
    if data.startswith("shop:"):
        shop_name = data[5:]
        await _process_gpay_with_name(msg, context, shop_name)
        return

    # ── "Type a new name" tapped — text handler will pick up what user types ──
    if data == "newshop":
        await msg.reply_text("✏️ Type the new shop name and send:")
        return

    # ── Compile now ──
    if data == "compile_now":
        records = context.user_data["records"]
        if not records:
            await msg.reply_text("📭 Nothing to compile yet — send some CSVs first.")
            return
        ok = await _do_compile(msg, context)
        if ok:
            await msg.reply_text(
                "Keep the current data or start fresh for the next batch?",
                reply_markup=_post_compile_keyboard(),
            )
        return

    # ── Wait for more files ──
    if data == "wait_more":
        await msg.reply_text(
            "👍 Got it — send more CSVs whenever you're ready.\n"
            "Use /compile when you want the report."
        )
        return

    # ── Keep data after compile ──
    if data == "keep_data":
        n = len(context.user_data["records"])
        await msg.reply_text(
            f"📁 Data kept ({n} records).  Send more CSVs or /compile again anytime."
        )
        return

    # ── Reset / clear confirmed (used by both /reset and post-compile prompt) ──
    if data == "reset_yes":
        context.user_data["records"]          = []
        context.user_data["pending_csv"]      = None
        context.user_data["gpay_queue"]       = []
        context.user_data["asking_gpay_name"] = False
        context.user_data["seen_txn_ids"]     = set()
        # Keep known_shops — shop names stay useful across sessions
        await msg.reply_text(
            "🔄 All records cleared!  Shop name history is kept.\n"
            "Send fresh CSVs whenever you're ready."
        )
        return

    # ── Reset cancelled ──
    if data == "reset_no":
        await msg.reply_text("↩️ Reset cancelled — your data is safe.")
        return

    logger.warning("Unknown callback data: %s", data)


# ══════════════════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════

def main() -> None:
    if not BOT_TOKEN:
        raise RuntimeError(
            "BOT_TOKEN is not set.  "
            "Copy .env.example → .env and add your token."
        )

    data_dir = Path(os.getenv("DATA_DIR", "data"))
    data_dir.mkdir(exist_ok=True)
    persistence = PicklePersistence(filepath=str(data_dir / "bot_state.pkl"))

    app = (
        Application.builder()
        .token(BOT_TOKEN)
        .persistence(persistence)
        .build()
    )

    app.add_handler(CommandHandler("start",   cmd_start))
    app.add_handler(CommandHandler("help",    cmd_start))
    app.add_handler(CommandHandler("status",  cmd_status))
    app.add_handler(CommandHandler("compile", cmd_compile))
    app.add_handler(CommandHandler("reset",   cmd_reset))
    app.add_handler(CommandHandler("debug",   cmd_debug))
    app.add_handler(CallbackQueryHandler(handle_callback))
    app.add_handler(MessageHandler(filters.Document.ALL, handle_document))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text))

    logger.info("Bot is running — press Ctrl+C to stop.")
    app.run_polling(drop_pending_updates=True)


if __name__ == "__main__":
    main()
